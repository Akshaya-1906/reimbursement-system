from flask import Flask, render_template, request, redirect, session, send_from_directory
import sqlite3, os, json
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from flask import send_file
import io
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle,
    Spacer, Image
)
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
from reportlab.lib.units import mm
from PyPDF2 import PdfMerger
import os


import pdfkit
from flask import make_response

import inflect
import io


app = Flask(__name__)
app.secret_key = "secret123"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "database.db")

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
BILL_FOLDER = os.path.join(UPLOAD_FOLDER, "bills")
PASSBOOK_FOLDER = os.path.join(UPLOAD_FOLDER, "passbooks")

os.makedirs(BILL_FOLDER, exist_ok=True)
os.makedirs(PASSBOOK_FOLDER, exist_ok=True)

def init_db():
    db = get_db()
    cur = db.cursor()

    # Admin table
    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        username TEXT PRIMARY KEY,
        password TEXT NOT NULL
    )
    """)

    # Student club passwords
    cur.execute("""
    CREATE TABLE IF NOT EXISTS student_passwords (
        club_name TEXT PRIMARY KEY,
        password TEXT NOT NULL
    )
    """)

    # Requests table (THIS FIXES YOUR ERROR)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        bill TEXT,
        passbook TEXT,
        table_data TEXT,
        status TEXT DEFAULT 'PENDING'
    )
    """)

    db.commit()

# ---------------- DB CONNECTION ----------------
def get_db():
    return sqlite3.connect(DB_PATH)



def number_to_words(n):
    
    p = inflect.engine()
    return p.number_to_words(n).replace("-", " ").title()

# ---------------- LOGIN ----------------
@app.route("/")
def home():
    return render_template("login.html")


@app.route("/login", methods=["POST"])
def login():
    role = request.form["role"]
    userid = request.form["userid"]
    password = request.form["password"]

    db = get_db()
    cur = db.cursor()

    # ---------- ADMIN LOGIN ----------
    if role == "admin":
        if userid != "admin":
            return render_template(
                "login.html",
                error="Invalid admin user ID"
            )

        cur.execute("SELECT password FROM users WHERE username='admin'")
        row = cur.fetchone()

        if row and row[0] == password:
            session["role"] = "admin"
            session["user"] = "admin"
            return redirect("/admin")
        else:
            return render_template(
                "login.html",
                error="Invalid admin password"
            )

    # ---------- STUDENT LOGIN ----------
    if role == "student":
        if userid != "student":
            return render_template(
                "login.html",
                error="Invalid student user ID"
            )

        cur.execute(
            "SELECT club_name FROM student_passwords WHERE password=?",
            (password,)
        )
        row = cur.fetchone()

        if row:
            session["role"] = "student"
            session["user"] = row[0]
            return redirect("/student")
        else:
            return render_template(
                "login.html",
                error="Invalid student password"
            )

    return render_template(
        "login.html",
        error="Invalid login attempt"
    )


# ---------------- STUDENT ----------------
@app.route("/student")
def student():
    if session.get("role") != "student":
        return redirect("/")
    return render_template("student.html")

@app.route("/submit", methods=["POST"])
def submit():

    if session.get("role") != "student":
        return redirect("/")

    # 🔹 MULTIPLE BILLS
    bills = request.files.getlist("bills")
    if not bills or bills[0].filename == "":
        return "No bills uploaded", 400

    bill_paths = []
    for bill in bills:
        filename = secure_filename(bill.filename)
        path = f"uploads/bills/{filename}"
        bill.save(os.path.join(BASE_DIR, path))
        bill_paths.append(path)

    # 🔹 PASSBOOK
    passbook = request.files["passbook"]
    pb_name = secure_filename(passbook.filename)
    pb_path = f"uploads/passbooks/{pb_name}"
    passbook.save(os.path.join(BASE_DIR, pb_path))

    # 🔹 TABLE DATA
    table_data = request.form["tableData"]

    # 🔹 SAVE TO DB
    db = get_db()
    cur = db.cursor()
    cur.execute("""
        INSERT INTO requests (student_id, bill, passbook, table_data, status)
        VALUES (?, ?, ?, ?, ?)
    """, (
        session["user"],
        json.dumps(bill_paths),   # IMPORTANT: list of bills
        pb_path,
        table_data,
        "PENDING"
    ))
    db.commit()
    db.close()

    return render_template("success.html")


@app.route("/student/status")
def student_status():
    if session.get("role") != "student":
        return redirect("/")

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT id, status FROM requests WHERE student_id=?",
        (session["user"],)
    )
    requests = cur.fetchall()
    db.close()

    return render_template("student_status.html", requests=requests)

@app.route("/student/download/<int:req_id>")
def student_download(req_id):
    if session.get("role") != "student":
        return redirect("/")

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT table_data FROM requests WHERE id=?", (req_id,))
    data = json.loads(cur.fetchone()[0])
    db.close()

    return generate_excel(data, f"student_request_{req_id}.xlsx")




    db = get_db()
    cur = db.cursor()
    cur.execute("""
        INSERT INTO requests (student_id, bill, passbook, table_data, status)
        VALUES (?, ?, ?, ?, ?)
    """, (
        session["user"],
        bill_path,
        passbook_path,
        table_data,
        "PENDING"
    ))
    db.commit()
    db.close()

    return "Submitted Successfully! <br><a href='/student'>Go Back</a>"

# ---------------- ADMIN ----------------
@app.route("/admin")
def admin():
    if session.get("role") != "admin":
        return redirect("/")

    db = get_db()
    cur = db.cursor()

    # All requests
    cur.execute("SELECT * FROM requests")
    data = cur.fetchall()

    # ✅ Approved request count (auto-updates after delete/approve)
    cur.execute("SELECT COUNT(*) FROM requests WHERE status = 'APPROVED'")
    approved_count = cur.fetchone()[0]

    db.close()

    return render_template(
        "admin.html",
        data=data,
        approved_count=approved_count
    )


# ---------------- CONSOLIDATED VIEW ----------------
@app.route("/admin/request/<int:req_id>")
def admin_view(req_id):
    if session.get("role") != "admin":
        return redirect("/")

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM requests WHERE id=?", (req_id,))
    r = cur.fetchone()
    db.close()

    table_json = json.loads(r[4])   # FULL table data
    bill_list = json.loads(r[2])    # list of bill PDFs

    return render_template(
        "admin_view.html",
        request_data=r,
        bills=bill_list,
        table_data=table_json
    )


# ---------------- APPROVE / DISAPPROVE ----------------
@app.route("/update/<int:req_id>/<status>")
def update(req_id, status):
    if session.get("role") != "admin":
        return redirect("/")

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "UPDATE requests SET status=? WHERE id=?",
        (status, req_id)
    )
    db.commit()
    db.close()

    return redirect("/admin")

@app.route("/admin/download/<int:req_id>")
def admin_download(req_id):
    if session.get("role") != "admin":
        return redirect("/")

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT table_data FROM requests WHERE id=?", (req_id,))
    data = json.loads(cur.fetchone()[0])
    db.close()

    return generate_excel(data, f"admin_request_{req_id}.xlsx")

def generate_excel(table_json, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reimbursement"

    headers = [
        "S. No",
        "Description",
        "Bill Amount",
        "To Pay Amount",
        "Name & Bank Acc. No.",
        "IFSC Code",
        "Branch"
    ]
    ws.append(headers)

    rows = table_json.get("rows", [])
    if not rows:
        return

    start_row = 2  # first data row
    for r in rows:
        ws.append(r)

    end_row = start_row + len(rows) - 1

    # 🔥 MERGE COLUMNS (VERTICALLY)
    ws.merge_cells(start_row=start_row, start_column=4,
                   end_row=end_row, end_column=4)

    ws.merge_cells(start_row=start_row, start_column=5,
                   end_row=end_row, end_column=5)

    ws.merge_cells(start_row=start_row, start_column=6,
                   end_row=end_row, end_column=6)

    ws.merge_cells(start_row=start_row, start_column=7,
                   end_row=end_row, end_column=7)

    # Center align merged cells
    from openpyxl.styles import Alignment
    for col in [4, 5, 6, 7]:
        ws.cell(row=start_row, column=col).alignment = Alignment(
            vertical="center", horizontal="center"
        )

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/admin/generate-letter", methods=["POST"])
def generate_letter():
    if session.get("role") != "admin":
        return redirect("/")

    subject = request.form["subject"]
    date_raw = request.form["letter_date"]
    y, m, d = date_raw.split("-")
    letter_date = f"{d}-{m}-{y}"

    request_ids = request.form.getlist("request_ids")

    db = get_db()
    cur = db.cursor()

    grouped_data = []   # 🔥 each request stays separate
    grand_total = 0

    for rid in request_ids:
        cur.execute(
            "SELECT table_data FROM requests WHERE id=?",
            (rid,)
        )
        table_json = json.loads(cur.fetchone()[0])

        rows = table_json["rows"]

        to_pay = rows[0][3]
        bank = rows[0][4]
        ifsc = rows[0][5]
        branch = rows[0][6]

        subtotal = sum(float(r[2]) for r in rows)
        grand_total += subtotal

        grouped_data.append({
            "rows": rows,
            "to_pay": to_pay,
            "bank": bank,
            "ifsc": ifsc,
            "branch": branch,
            "subtotal": subtotal
        })

    db.close()

    session["letter_data"] = {
    "subject": subject,
    "letter_date": letter_date,
    "groups": grouped_data,
    "total": grand_total,
    "total_words": number_to_words(int(grand_total))
}
    session["letter_subject"] = subject
    session["letter_date"] = letter_date
    session["letter_groups"] = grouped_data
    session["letter_total"] = grand_total
    session["letter_total_words"] = number_to_words(int(grand_total))


    
    return render_template(
    "letter.html",
    subject=subject,
    letter_date=letter_date,
    groups=grouped_data,
    total=grand_total,
    total_words=number_to_words(int(grand_total))
)

def draw_first_page_header(canvas, doc, letter_date):
    canvas.saveState()

    top_y = doc.pagesize[1]

    # Logo
    canvas.drawImage(
        "static/images/psg_logo.png",
        doc.leftMargin,
        top_y - 70,
        width=45,
        height=45,
        preserveAspectRatio=True
    )

    # Institute Name
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawCentredString(
        doc.pagesize[0] / 2,
        top_y - 45,
        "PSG INSTITUTE OF TECHNOLOGY AND APPLIED RESEARCH"
    )

    # Student Council
    canvas.setFont("Helvetica-Bold", 11)
    canvas.drawCentredString(
        doc.pagesize[0] / 2,
        top_y - 60,
        "STUDENT COUNCIL"
    )

    # ✅ Date — ONLY FIRST PAGE
    canvas.setFont("Helvetica", 10)
    canvas.drawRightString(
        doc.pagesize[0] - doc.rightMargin,
        top_y - 80,
        f"Date: {letter_date}"
    )

    canvas.restoreState()
def draw_later_pages_header(canvas, doc):
    canvas.saveState()

    top_y = doc.pagesize[1]

    # Logo
    canvas.drawImage(
        "static/images/psg_logo.png",
        doc.leftMargin,
        top_y - 70,
        width=45,
        height=45,
        preserveAspectRatio=True
    )

    # Institute Name
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawCentredString(
        doc.pagesize[0] / 2,
        top_y - 45,
        "PSG INSTITUTE OF TECHNOLOGY AND APPLIED RESEARCH"
    )

    # Student Council
    canvas.setFont("Helvetica-Bold", 11)
    canvas.drawCentredString(
        doc.pagesize[0] / 2,
        top_y - 60,
        "STUDENT COUNCIL"
    )

    canvas.restoreState()



# =========================
# FOOTER (LAST PAGE ONLY)
# =========================
class LastPageFooterCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            if self._pageNumber == total_pages:
                self.draw_footer()
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_footer(self):
        y = 25 * mm
        self.setFont("Helvetica", 10)

        self.drawString(
            40,
            y,
            "Faculty Advisor – Student Council"
        )

        self.drawRightString(
            self._pagesize[0] - 40,
            y,
            "Principal"
        )


# =========================
# MAIN PDF FUNCTION
# =========================
def generate_letter_pdf(subject, letter_date, groups, total, total_words):
    buffer = io.BytesIO()

    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=40,
        rightMargin=40,
        topMargin=90,     # space for header
        bottomMargin=40
    )

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height,
        id="normal"
    )

    doc.addPageTemplates([
        doc.addPageTemplates([
         PageTemplate(
            id="FirstPage",
            frames=[frame],
            onPage=lambda c, d: draw_first_page_header(c, d, letter_date)
         ),
         PageTemplate(
            id="LaterPages",
            frames=[frame],
            onPage=draw_later_pages_header
         )
        ])

    ])

    styles = getSampleStyleSheet()
    content = []

    # ================= BODY TEXT =================
    content.append(Paragraph("<b>Submitted to Principal:</b>", styles["Normal"]))
    content.append(Spacer(1, 6))
    content.append(Paragraph(f"<b>Sub:</b> {subject}", styles["Normal"]))
    content.append(Spacer(1, 10))

    content.append(Paragraph(
        "Principal’s kind permission is requested to settle the list of bills given below. "
        "Photocopy of the bills are enclosed for your kind perusal.",
        styles["Normal"]
    ))

    content.append(Spacer(1, 15))

    # ================= TABLE STYLES =================
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
        alignment=TA_CENTER
    )

    header_style = ParagraphStyle(
        "header",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold"
    )

    # ================= TABLE DATA =================
    table_data = [[
        Paragraph("S. No", header_style),
        Paragraph("Description", header_style),
        Paragraph("Bill Amount", header_style),
        Paragraph("To Pay Amount", header_style),
        Paragraph("Name & Bank Acc. No.", header_style),
        Paragraph("IFSC Code", header_style),
        Paragraph("Branch", header_style),
    ]]

    style_cmds = [
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (2,1), (-1,-1), "CENTER"),
    ]

    row_index = 1
    sno = 1

    for g in groups:
        start = row_index
        first = True

        for r in g["rows"]:
            table_data.append([
                Paragraph(str(sno), cell_style),
                Paragraph(str(r[1]), cell_style),
                Paragraph(str(r[2]), cell_style),
                Paragraph(str(g["to_pay"]) if first else "", cell_style),
                Paragraph(str(g["bank"]) if first else "", cell_style),
                Paragraph(str(g["ifsc"]) if first else "", cell_style),
                Paragraph(str(g["branch"]) if first else "", cell_style),
            ])
            first = False
            row_index += 1

        end = row_index - 1

        style_cmds += [
            ("SPAN", (0, start), (0, end)),
            ("SPAN", (3, start), (3, end)),
            ("SPAN", (4, start), (4, end)),
            ("SPAN", (5, start), (5, end)),
            ("SPAN", (6, start), (6, end)),
        ]

        sno += 1

    # ================= TOTAL ROW =================
    table_data.append([
        Paragraph("", cell_style),
        Paragraph("Total", header_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph(f"{total} ({total_words} Only)", header_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
    ])

    style_cmds += [
        ("SPAN", (0, row_index), (1, row_index)),
        ("SPAN", (4, row_index), (6, row_index)),
        ("ALIGN", (4, row_index), (6, row_index), "CENTER"),
    ]

    table = Table(
        table_data,
        colWidths=[35, 110, 65, 70, 140, 55, 40],
        repeatRows=1
    )

    table.setStyle(TableStyle(style_cmds))
    content.append(table)

    # ================= BUILD =================
    doc.build(
        content,
        canvasmaker=LastPageFooterCanvas
    )

    buffer.seek(0)
    return buffer


@app.route("/admin/letter/download")
def download_letter():
    if session.get("role") != "admin":
        return redirect("/")

    buffer = generate_letter_pdf(
        session["letter_subject"],
        session["letter_date"],
        session["letter_groups"],
        session["letter_total"],
        session["letter_total_words"]
    )

    return send_file(
        buffer,
        as_attachment=True,
        download_name="Official_Letter.pdf",
        mimetype="application/pdf"
    )
def merge_pdfs(pdf_paths, output_buffer):
    merger = PdfMerger()

    for path in pdf_paths:
        if os.path.exists(path):
            merger.append(path)

    merger.write(output_buffer)
    merger.close()
from PyPDF2 import PdfMerger
from flask import send_file
import io, json, os

@app.route("/admin/download-approved-bills")
def download_approved_bills():
    if session.get("role") != "admin":
        return redirect("/")

    db = get_db()
    cur = db.cursor()

    cur.execute("""
        SELECT bill FROM requests WHERE status='APPROVED'
    """)
    rows = cur.fetchall()
    db.close()

    merger = PdfMerger()

    for row in rows:
        bill_list = json.loads(row[0])   # bill is stored as JSON list
        for bill_path in bill_list:
            full_path = os.path.join(BASE_DIR, bill_path)
            if os.path.exists(full_path):
                merger.append(full_path)

    # 🔥 WRITE INTO BYTESIO
    output = io.BytesIO()
    merger.write(output)
    merger.close()

    # 🔥 THIS IS CRITICAL
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Approved_Bills.pdf",
        mimetype="application/pdf"
    )


@app.route("/admin/download-approved-passbooks")
def download_approved_passbooks():
    if session.get("role") != "admin":
        return redirect("/")

    db = get_db()
    cur = db.cursor()

    cur.execute("""
        SELECT passbook FROM requests WHERE status='APPROVED'
    """)
    rows = cur.fetchall()
    db.close()

    merger = PdfMerger()

    for row in rows:
        full_path = os.path.join(BASE_DIR, row[0])
        if os.path.exists(full_path):
            merger.append(full_path)

    output = io.BytesIO()
    merger.write(output)
    merger.close()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Approved_Passbooks.pdf",
        mimetype="application/pdf"
    )
@app.route("/admin/delete-selected", methods=["POST"])
def delete_selected():
    if session.get("role") != "admin":
        return redirect("/")

    ids = request.form.getlist("delete_ids")

    if ids:
        cursor = sqlite3.connect("database.db").cursor()
        cursor.execute(
            f"DELETE FROM requests WHERE id IN ({','.join(['?']*len(ids))})",
            ids
        )
        cursor.connection.commit()

    return redirect("/admin")

# ---------------- FILE SERVING ----------------
@app.route("/files/<path:filepath>")
def serve_file(filepath):
    return send_from_directory(BASE_DIR, filepath)


# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

with app.app_context():
    init_db()

# ---------------- RUN ----------------
if __name__ == "__main__":
    app.run(debug=True)
